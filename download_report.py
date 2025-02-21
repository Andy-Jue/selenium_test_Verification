from selenium import webdriver
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import random
import logging
from response import catch_response 
from selenium.webdriver.support.ui import Select
import string
import pandas as pd
from openpyxl import load_workbook

file_path = "C:\\Users\\iou85\\Downloads\\清冊640.xlsx"
spreadsheet = pd.ExcelFile(file_path)

# # 檢查工作表名稱
print(spreadsheet.sheet_names)
df = pd.read_excel(file_path, sheet_name="GHG inventory report",skiprows=[0])
print(df.head())
df.to_excel('output1.xlsx', index=False)

file_path = "C:\\Users\\iou85\\Desktop\\selenium測試(固定)\\output1.xlsx"
spreadsheet = pd.ExcelFile(file_path)

# # 檢查工作表名稱
print(spreadsheet.sheet_names)
df = pd.read_excel(file_path,skiprows=[0])
print(df)



# wb = load_workbook(file_path)
# ws = wb.active

# data = []
# for row in ws.iter_rows(values_only=True):
#     data.append(row)

# df = pd.DataFrame(data)
# print(df)
# df.to_excel('output.xlsx', index=False)